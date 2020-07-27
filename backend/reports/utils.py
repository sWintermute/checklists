from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import Color
from tempfile import NamedTemporaryFile

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell import Cell


def create_file(query, questions, host):
    wb = Workbook()
    stream = None
    ws = wb.active

    # HEADER
    header, questions_id = create_header(questions)
    ws.append(header)
    row = ws.row_dimensions[1]
    c = Color(indexed=22)
    f = PatternFill(fill_type='darkGrid', start_color=c, end_color=c)
    a = Alignment(horizontal='center', vertical='center', text_rotation=0,
                  wrap_text=True, shrink_to_fit=False, indent=2)
    row.fill = f
    row.alignment = a
    row.height = 50
    end_cell = Cell(worksheet=ws, column=len(header), row=5)
    tab = Table(displayName="Table1", ref="A1:" + end_cell.coordinate)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                           showLastColumn=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # BODY
    body = get_data_from_query_as_lists(query, questions_id, host)
    for data in body:
        ws.append(data)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        return stream


def create_header(questions):
    result = ['Ссылка', 'Номер ответа', 'Дата создания', 'Почта']
    questions_id_list = []
    for question in questions:
        result.append(question.text)
        questions_id_list.append(question.id)
    return result, questions_id_list


def get_data_from_query_as_lists(query, questions_id, host):
    results = list()
    LINK = f"{host}"
    for q in query:
        new_piece = [
            f'{LINK}{q.id}',
            q.id,
            q.created,
            q.user.email
        ]
        answers = q.answers.all()
        for q_id in questions_id:
            item = '-'
            for answer in answers:
                if answer.question.id == q_id:
                    item = float(answer.body) if answer.body.isdigit() \
                            else answer.body
                    break
            new_piece.append(item)
        results.append(new_piece)
    return results
